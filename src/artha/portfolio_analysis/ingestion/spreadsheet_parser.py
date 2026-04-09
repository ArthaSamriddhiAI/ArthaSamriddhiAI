"""Parse .xlsx portfolio files into the canonical portfolio JSON structure."""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

REQUIRED_COLUMNS = {
    "instrument_name",
    "isin_or_cin",
    "asset_class",
    "current_value_inr",
    "purchase_date",
    "purchase_price_per_unit",
    "quantity_or_units",
}
OPTIONAL_COLUMNS = {"folio_or_account_no"}

VALID_ASSET_CLASSES = {
    "listed_equity",
    "mutual_fund",
    "pms",
    "aif_cat1",
    "aif_cat2",
    "aif_cat3",
    "unlisted_equity",
    "cash",
}


def _normalise_header(header: str) -> str:
    """Lowercase, strip, replace spaces/dashes with underscores."""
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def _parse_date(val: Any) -> date | None:
    """Best-effort date coercion."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_spreadsheet(file_bytes: bytes) -> dict:
    """Parse an .xlsx file into the canonical portfolio JSON structure.

    Returns
    -------
    dict with keys: holdings, asset_class_breakdown, data_quality_summary
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Spreadsheet contains no active worksheet")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Spreadsheet must contain a header row and at least one data row")

    # Map column indices
    raw_headers = [str(c).strip() if c else "" for c in rows[0]]
    headers = [_normalise_header(h) for h in raw_headers]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    col_idx: dict[str, int] = {h: i for i, h in enumerate(headers) if h}

    today = date.today()
    holdings: list[dict] = []
    total_value = 0.0
    data_gaps: list[dict] = []
    asset_class_totals: dict[str, float] = {}

    for row_num, row in enumerate(rows[1:], start=1):
        def _cell(name: str) -> Any:
            idx = col_idx.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        instrument_name = _cell("instrument_name")
        if not instrument_name:
            continue  # skip blank rows

        holding_id = f"h-{row_num:03d}"
        asset_class = str(_cell("asset_class") or "").strip().lower()
        current_value = _safe_float(_cell("current_value_inr"))
        purchase_date = _parse_date(_cell("purchase_date"))
        purchase_price = _safe_float(_cell("purchase_price_per_unit"))
        quantity = _safe_float(_cell("quantity_or_units"))
        isin_or_cin = _cell("isin_or_cin")
        folio = _cell("folio_or_account_no")

        # Validate asset class
        if asset_class not in VALID_ASSET_CLASSES:
            data_gaps.append({
                "holding_id": holding_id,
                "field": "asset_class",
                "issue": f"Invalid asset_class '{asset_class}'",
            })
            asset_class = "cash"  # fallback

        # Compute derived fields
        gaps_for_holding: list[str] = []

        if current_value is None:
            gaps_for_holding.append("current_value_inr")
            current_value = 0.0
        if purchase_date is None:
            gaps_for_holding.append("purchase_date")
        if purchase_price is None:
            gaps_for_holding.append("purchase_price_per_unit")
        if quantity is None:
            gaps_for_holding.append("quantity_or_units")

        holding_period_days: int | None = None
        if purchase_date is not None:
            holding_period_days = (today - purchase_date).days

        ltcg_eligible: bool | None = None
        if holding_period_days is not None:
            # Simplified threshold: 365 for listed equity/MF, 730 for unlisted, 1095 for AIF
            threshold_map = {
                "listed_equity": 365,
                "mutual_fund": 365,
                "pms": 365,
                "aif_cat1": 1095,
                "aif_cat2": 1095,
                "aif_cat3": 365,
                "unlisted_equity": 730,
                "cash": 0,
            }
            threshold = threshold_map.get(asset_class, 365)
            ltcg_eligible = holding_period_days >= threshold

        cost_basis: float | None = None
        if purchase_price is not None and quantity is not None:
            cost_basis = purchase_price * quantity

        total_value += current_value
        asset_class_totals[asset_class] = asset_class_totals.get(asset_class, 0.0) + current_value

        if gaps_for_holding:
            data_gaps.append({
                "holding_id": holding_id,
                "field": ", ".join(gaps_for_holding),
                "issue": "missing_data",
            })

        holding = {
            "holding_id": holding_id,
            "instrument_name": str(instrument_name).strip(),
            "isin_or_cin": str(isin_or_cin).strip() if isin_or_cin else None,
            "asset_class": asset_class,
            "current_value_inr": current_value,
            "purchase_date": purchase_date.isoformat() if purchase_date else None,
            "purchase_price_per_unit": purchase_price,
            "quantity_or_units": quantity,
            "folio_or_account_no": str(folio).strip() if folio else None,
            "cost_basis": cost_basis,
            "weight_pct": 0.0,  # computed below
            "holding_period_days": holding_period_days,
            "ltcg_eligible": ltcg_eligible,
            "data_gaps": gaps_for_holding,
        }
        holdings.append(holding)

    # Compute weight_pct
    if total_value > 0:
        for h in holdings:
            h["weight_pct"] = round((h["current_value_inr"] / total_value) * 100, 4)

    # Asset class breakdown
    asset_class_breakdown = []
    for ac, val in sorted(asset_class_totals.items(), key=lambda x: -x[1]):
        pct = round((val / total_value) * 100, 2) if total_value > 0 else 0.0
        asset_class_breakdown.append({
            "asset_class": ac,
            "total_value_inr": val,
            "weight_pct": pct,
            "holdings_count": sum(1 for h in holdings if h["asset_class"] == ac),
        })

    data_quality_summary = {
        "total_holdings": len(holdings),
        "holdings_with_gaps": len({g["holding_id"] for g in data_gaps}),
        "total_data_gaps": len(data_gaps),
        "gap_details": data_gaps,
    }

    wb.close()

    return {
        "holdings": holdings,
        "asset_class_breakdown": asset_class_breakdown,
        "data_quality_summary": data_quality_summary,
        "total_value_inr": total_value,
    }
